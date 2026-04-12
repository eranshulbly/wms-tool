# -*- encoding: utf-8 -*-

from flask import request, send_file
from flask_restx import Resource, fields
from datetime import datetime
import pandas as pd
import openpyxl
import io

from .models import (
    TransportRoute, CustomerRouteMapping,
    DailyRouteManifest, CompanySchemaMapping
)
from .business.dealer_business import get_or_create_dealer
from .routes import rest_api, token_required, active_required

# --- Basic response models ---
basic_response = rest_api.model('EwayBasicResponse', {
    'success': fields.Boolean,
    'msg': fields.String
})

from functools import wraps
from .permissions import get_permissions

def eway_admin_required(f):
    @wraps(f)
    def decorator(*args, **kwargs):
        # Bug 36 fix: default-deny — if current_user is None, reject immediately.
        current_user = args[1] if len(args) > 1 else kwargs.get('current_user')
        if not current_user:
            return {'success': False, 'msg': 'Authentication required'}, 401
        perms = get_permissions(current_user.role)
        if not perms.get('eway_bill_admin'):
            return {'success': False, 'msg': 'E-way Bill Admin permission required'}, 403
        return f(*args, **kwargs)
    return decorator

def eway_filling_required(f):
    @wraps(f)
    def decorator(*args, **kwargs):
        # Bug 36 fix: default-deny — if current_user is None, reject immediately.
        current_user = args[1] if len(args) > 1 else kwargs.get('current_user')
        if not current_user:
            return {'success': False, 'msg': 'Authentication required'}, 401
        perms = get_permissions(current_user.role)
        if not (perms.get('eway_bill_admin') or perms.get('eway_bill_filling')):
            return {'success': False, 'msg': 'E-way Bill Filling permission required'}, 403
        return f(*args, **kwargs)
    return decorator

# ─────────────────────────────────────────────────────────────
# ROUTES
# ─────────────────────────────────────────────────────────────

@rest_api.route('/api/eway/routes')
class EwayRoutes(Resource):
    @token_required
    @active_required
    @eway_filling_required
    def get(self, current_user):
        try:
            routes = TransportRoute.get_all()
            return {'success': True, 'routes': [r.to_dict() for r in routes]}, 200
        except Exception as e:
            return {'success': False, 'msg': str(e), 'routes': []}, 400

    @token_required
    @active_required
    @eway_admin_required
    def post(self, current_user):
        """Create a route."""
        try:
            data = request.json
            route = TransportRoute(name=data['name'], description=data.get('description', ''))
            route.save()
            return {'success': True, 'msg': 'Route created', 'route_id': route.route_id}, 200
        except Exception as e:
            return {'success': False, 'msg': str(e)}, 400


# ─────────────────────────────────────────────────────────────
# CUSTOMER → ROUTE MAPPINGS
# ─────────────────────────────────────────────────────────────

@rest_api.route('/api/eway/customer-route-mappings')
class EwayCustomerRouteMappings(Resource):
    @token_required
    @active_required
    @eway_filling_required
    def get(self, current_user):
        """Get all mappings, or filter by route_id for popup."""
        try:
            route_id = request.args.get('route_id', type=int)
            if route_id:
                mappings = CustomerRouteMapping.get_for_route(route_id)
            else:
                mappings = CustomerRouteMapping.get_all()
            clean = [
                {k: v for k, v in m.items() if not hasattr(v, 'isoformat')}
                for m in mappings
            ]
            return {'success': True, 'mappings': clean}, 200
        except Exception as e:
            return {'success': False, 'msg': str(e)}, 400

    @token_required
    @active_required
    @eway_admin_required
    def post(self, current_user):
        """Add / update a single customer-route mapping (upsert by dealer)."""
        try:
            data = request.json
            dealer_code = data['customer_code'].strip()
            dealer_name = data.get('customer_name', '').strip() or dealer_code
            dealer_id = get_or_create_dealer(dealer_name=dealer_name, dealer_code=dealer_code)
            m = CustomerRouteMapping(
                dealer_id=dealer_id,
                route_id=int(data['route_id']),
                distance=int(data['distance'])
            )
            m.save()
            return {'success': True, 'msg': 'Mapping saved'}, 200
        except Exception as e:
            return {'success': False, 'msg': str(e)}, 400


@rest_api.route('/api/eway/customer-route-mappings/bulk')
class EwayCustomerRouteMappingsBulk(Resource):
    @token_required
    @active_required
    @eway_admin_required
    def post(self, current_user):
        """
        Bulk-import customer→route mappings from an Excel file.
        Expected columns (case-insensitive):
          Customer Code | Customer Name | Route Name | Distance (km)
        Route Name is matched case-insensitively to existing routes.
        """
        try:
            if 'file' not in request.files:
                return {'success': False, 'msg': 'No file uploaded'}, 400
            f = request.files['file']
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return {'success': False, 'msg': 'Empty file'}, 400

            headers = [str(h).strip().lower() if h else '' for h in rows[0]]

            def col(name):
                aliases = {
                    'customer code': ['customer code', 'customer_code', 'code', 'cust code'],
                    'customer name': ['customer name', 'customer_name', 'name', 'account name'],
                    'route name':   ['route name', 'route_name', 'route'],
                    'distance':     ['distance (km)', 'distance', 'dist km', 'dist', 'km'],
                }
                for a in aliases.get(name, [name]):
                    if a in headers:
                        return headers.index(a)
                return None

            ci_code  = col('customer code')
            ci_name  = col('customer name')
            ci_route = col('route name')
            ci_dist  = col('distance')

            if ci_code is None or ci_route is None or ci_dist is None:
                return {
                    'success': False,
                    'msg': 'Could not find required columns. Expected: Customer Code, Route Name, Distance (km)'
                }, 400

            # Pre-load all routes for case-insensitive name lookup
            all_routes = TransportRoute.get_all()
            route_map = {r.name.lower(): r.route_id for r in all_routes}

            imported = 0
            errors = []
            for i, row in enumerate(rows[1:], start=2):
                try:
                    code       = str(row[ci_code]).strip()  if row[ci_code]  else ''
                    route_name = str(row[ci_route]).strip() if row[ci_route] else ''
                    dist       = row[ci_dist]
                    if not code or not route_name or dist is None:
                        continue

                    route_id = route_map.get(route_name.lower())
                    if not route_id:
                        errors.append(f"Row {i}: Route '{route_name}' not found")
                        continue

                    cust_name = str(row[ci_name]).strip() if (ci_name is not None and row[ci_name]) else code
                    dealer_id = get_or_create_dealer(dealer_name=cust_name, dealer_code=code)
                    m = CustomerRouteMapping(
                        dealer_id=dealer_id,
                        route_id=route_id,
                        distance=int(float(str(dist)))
                    )
                    m.save()
                    imported += 1
                except Exception as row_err:
                    errors.append(f"Row {i}: {str(row_err)}")

            return {
                'success': True,
                'imported': imported,
                'errors': errors,
                'msg': f"Imported {imported} mappings" + (f" ({len(errors)} errors)" if errors else "")
            }, 200
        except Exception as e:
            return {'success': False, 'msg': str(e)}, 400


@rest_api.route('/api/eway/customer-route-mappings/template')
class EwayCustomerRouteMappingsTemplate(Resource):
    @token_required
    @active_required
    @eway_admin_required
    def get(self, current_user):
        """Download a sample Excel template for bulk upload."""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Customer Route Mapping"
            ws.append(['Customer Code', 'Customer Name', 'Route Name', 'Distance (km)'])
            ws.append(['HMC001', 'Hero Moto Corp - Delhi', 'UP Route A', 120])
            ws.append(['HMC002', 'Hero Moto Corp - UP', 'UP Route B', 85])
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 16

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return send_file(buf,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True,
                             download_name='customer_route_mapping_template.xlsx')
        except Exception as e:
            return {'success': False, 'msg': str(e)}, 400


@rest_api.route('/api/eway/customer-route-mappings/delete')
class EwayCustomerRouteMappingDelete(Resource):
    @token_required
    @active_required
    @eway_admin_required
    def post(self, current_user):
        """Remove a customer-route mapping by dealer code."""
        try:
            data = request.json
            CustomerRouteMapping.delete_by_dealer_code(data['customer_code'].strip())
            return {'success': True, 'msg': 'Customer removed from route'}, 200
        except Exception as e:
            return {'success': False, 'msg': str(e)}, 400


# ─────────────────────────────────────────────────────────────
# ROUTE CUSTOMERS  (admin-only: view & remove customers per route)
# ─────────────────────────────────────────────────────────────

@rest_api.route('/api/eway/route-customers/<int:route_id>')
class EwayRouteCustomers(Resource):
    @token_required
    @active_required
    @eway_admin_required
    def get(self, current_user, route_id):
        """List all customers assigned to a specific route. Admin only."""
        try:
            customers = CustomerRouteMapping.get_for_route(route_id)
            return {'success': True, 'customers': customers}, 200
        except Exception as e:
            return {'success': False, 'msg': str(e)}, 400


@rest_api.route('/api/eway/route-customers/remove')
class EwayRouteCustomerRemove(Resource):
    @token_required
    @active_required
    @eway_admin_required
    def post(self, current_user):
        """Remove a customer from their route by dealer code. Admin only."""
        try:
            data = request.json
            CustomerRouteMapping.delete_by_dealer_code(data['customer_code'].strip())
            return {'success': True, 'msg': 'Customer removed from route'}, 200
        except Exception as e:
            return {'success': False, 'msg': str(e)}, 400


# ─────────────────────────────────────────────────────────────
# DAILY MANIFEST
# ─────────────────────────────────────────────────────────────

@rest_api.route('/api/eway/manifest')
class EwayManifest(Resource):
    @token_required
    @active_required
    @eway_filling_required
    def get(self, current_user):
        try:
            date_str = request.args.get('date', datetime.utcnow().strftime('%Y-%m-%d'))
            manifests = DailyRouteManifest.get_for_date(date_str)
            enriched = []
            for m in (manifests or []):
                customers = CustomerRouteMapping.get_for_route(m['route_id'])
                enriched.append({**m, 'customer_count': len(customers)})
            return {'success': True, 'manifests': enriched}, 200
        except Exception as e:
            return {'success': False, 'msg': str(e)}, 400

    @token_required
    @active_required
    @eway_filling_required
    def post(self, current_user):
        try:
            data = request.json
            date_str = data.get('date', datetime.utcnow().strftime('%Y-%m-%d'))
            for asg in data.get('assignments', []):
                DailyRouteManifest(
                    route_id=asg['route_id'],
                    vehicle_number=asg['vehicle_number'],
                    manifest_date=date_str
                ).save()
            return {'success': True, 'msg': 'Daily manifest saved'}, 200
        except Exception as e:
            return {'success': False, 'msg': str(e)}, 400


# ─────────────────────────────────────────────────────────────
# CSV UPLOAD  — enriches rows via Customer → Route → Vehicle
# ─────────────────────────────────────────────────────────────

@rest_api.route('/api/eway/upload')
class EwayUpload(Resource):
    @token_required
    @active_required
    @eway_filling_required
    def post(self, current_user):
        try:
            if 'file' not in request.files:
                return {'success': False, 'msg': 'No file uploaded'}, 400
            file = request.files['file']
            if not file.filename:
                return {'success': False, 'msg': 'No file selected'}, 400

            company_id = request.form.get('company_id', type=int)
            if not company_id:
                return {'success': False, 'msg': 'company_id is required'}, 400

            schema = CompanySchemaMapping.get_for_company(company_id)
            if not schema:
                return {'success': False, 'msg': 'No schema found for this company. Configure it in Company Schema Config tab.'}, 400

            # Parse CSV
            raw = file.read()
            sep = '\t' if b'\t' in raw[:500] else ','
            df = pd.read_csv(io.BytesIO(raw), sep=sep)

            irn_col     = schema['irn_col']
            ccode_col   = schema['customer_code_col']
            cname_col   = schema['customer_name_col']
            invoice_col = schema['invoice_no_col']

            missing = [c for c in [irn_col, ccode_col, cname_col, invoice_col] if c not in df.columns]
            if missing:
                return {'success': False, 'msg': f"CSV columns not found: {', '.join(missing)}"}, 400

            today = datetime.utcnow().strftime('%Y-%m-%d')
            results = []

            for _, row in df.iterrows():
                irn = str(row.get(irn_col, '') or '').strip()
                if not irn or irn.lower() == 'nan':
                    continue  # Only rows with an IRN are e-invoices

                c_code  = str(row.get(ccode_col,   '') or '').strip()
                c_name  = str(row.get(cname_col,   '') or '').strip()
                inv_no  = str(row.get(invoice_col, '') or '').strip()

                route_id   = None
                vehicle_no = None
                distance   = None

                # Lookup: Customer Code (dealer_code) → route + distance
                cm = CustomerRouteMapping.find_by_dealer_code(c_code)
                if cm:
                    route_id = cm.route_id
                    distance = cm.distance
                    if route_id:
                        # Lookup: route + date → vehicle
                        vehicle_no = DailyRouteManifest.get_vehicle_for_route_date(route_id, today)

                status = 'Complete' if (distance and vehicle_no) else 'Incomplete'
                results.append({
                    'invoice_no':    inv_no,
                    'customer_code': c_code,
                    'customer_name': c_name,
                    'irn':           irn,
                    'distance':      distance,
                    'vehicle_no':    vehicle_no,
                    'route_id':      route_id,
                    'status':        status
                })

            # ── Extract date from filename ────────────────────────────────────
            import re as _re
            file_date = None
            fname = getattr(file, 'filename', '') or ''
            for pattern, fmt in [
                (r'(\d{2}-\d{2}-\d{4})', 'dmy'),   # 24-03-2026
                (r'(\d{2}_\d{2}_\d{4})', 'dmy'),   # 24_03_2026
                (r'(\d{4}-\d{2}-\d{2})', 'ymd'),   # 2026-03-24
            ]:
                m = _re.search(pattern, fname)
                if m:
                    try:
                        s = m.group(1).replace('_', '-')
                        parts = s.split('-')
                        if fmt == 'dmy':
                            file_date = f"{parts[2]}-{parts[1]}-{parts[0]}"
                        else:
                            file_date = s
                    except Exception:
                        pass
                    break

            return {
                'success': True,
                'data': results,
                'meta': {
                    'filename': fname,
                    'row_count': len(results),
                    'file_date': file_date,
                    'today': today
                }
            }, 200
        except Exception as e:
            return {'success': False, 'msg': str(e)}, 400


# ─────────────────────────────────────────────────────────────
# JSON GENERATION
# ─────────────────────────────────────────────────────────────

@rest_api.route('/api/eway/generate-json')
class EwayGenerateJson(Resource):
    @token_required
    @active_required
    @eway_filling_required
    def post(self, current_user):
        try:
            data = request.json
            payload = []
            for item in data:
                if not item.get('irn') or not item.get('vehicle_no') or not item.get('distance'):
                    return {'success': False, 'msg': 'Missing IRN, Vehicle No, or Distance in one or more rows.'}, 400
                payload.append({
                    "Irn":           item['irn'],
                    "Distance":      int(item['distance']),
                    "TransporterId": "",
                    "TransDocNo":    "",
                    "TransDocDate":  "",
                    "VehicleNo":     item['vehicle_no'],
                    "VehType":       "R"
                })
            return {'success': True, 'json_payload': payload, 'msg': f'Generated {len(payload)} records.'}, 200
        except Exception as e:
            return {'success': False, 'msg': str(e)}, 400


# ─────────────────────────────────────────────────────────────
# SCHEMA MAPPINGS  (company CSV column config)
# ─────────────────────────────────────────────────────────────

@rest_api.route('/api/eway/schema-mappings')
class EwaySchemaMappings(Resource):
    @token_required
    @active_required
    @eway_filling_required
    def get(self, current_user):
        try:
            company_id = request.args.get('company_id', type=int)
            if not company_id:
                return {'success': False, 'msg': 'company_id is required'}, 400
            schema = CompanySchemaMapping.get_for_company(company_id)
            return {'success': True, 'schema': schema}, 200
        except Exception as e:
            return {'success': False, 'msg': str(e)}, 400

    @token_required
    @active_required
    @eway_filling_required
    def post(self, current_user):
        try:
            data = request.json
            schema = CompanySchemaMapping(
                company_id=data['company_id'],
                invoice_no_col=data['invoice_no_col'],
                customer_code_col=data['customer_code_col'],
                customer_name_col=data['customer_name_col'],
                irn_col=data['irn_col'],
                amount_col=data.get('amount_col', '')
            )
            schema.save()
            return {'success': True, 'msg': 'Schema saved'}, 200
        except Exception as e:
            return {'success': False, 'msg': str(e)}, 400
