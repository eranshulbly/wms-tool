# -*- encoding: utf-8 -*-
"""
Invoice management routes:
  POST /api/invoices/upload
  POST /api/invoices/download-errors
  GET  /api/invoices/statistics
  GET  /api/invoices
  GET  /api/invoices/<invoice_id>
  GET  /api/invoices/supply-sheet/download
"""

from datetime import datetime

import werkzeug
from flask import request, make_response, send_file
from flask_restx import Resource, fields, reqparse

from ..extensions import rest_api
from ..core.auth import token_required, active_required, upload_permission_required
from ..models import Invoice, PotentialOrder, Warehouse, Company, mysql_manager
from ..db_manager import partition_filter
from ..services import invoice_service
from ..core.logging import get_logger

logger = get_logger(__name__)

# ── Models ───────────────────────────────────────────────────────────────────

invoice_upload_parser = reqparse.RequestParser()
invoice_upload_parser.add_argument('file',
                                   type=werkzeug.datastructures.FileStorage,
                                   location='files',
                                   required=True,
                                   help='Excel/CSV file with invoice data')
invoice_upload_parser.add_argument('warehouse_id',
                                   type=int, location='form', required=True,
                                   help='Warehouse ID must be provided')
invoice_upload_parser.add_argument('company_id',
                                   type=int, location='form', required=True,
                                   help='Company ID must be provided')

invoice_upload_response = rest_api.model('InvoiceUploadResponse', {
    'success':            fields.Boolean(description='Success status of upload'),
    'msg':                fields.String(description='Message describing the result'),
    'invoices_processed': fields.Integer(description='Number of invoices processed'),
    'orders_completed':   fields.Integer(description='Number of orders completed'),
    'errors':             fields.List(fields.String, description='List of errors encountered'),
    'upload_batch_id':    fields.String(description='Batch ID for tracking'),
    'has_errors':         fields.Boolean(description='Whether errors occurred'),
})

invoice_statistics_response = rest_api.model('InvoiceStatisticsResponse', {
    'success':        fields.Boolean(description='Success status'),
    'total_invoices': fields.Integer(description='Total number of invoices'),
    'unique_orders':  fields.Integer(description='Number of unique orders'),
    'total_amount':   fields.Float(description='Total invoice amount'),
    'recent_batches': fields.List(fields.Raw, description='Recent upload batches'),
})

invoice_model = rest_api.model('Invoice', {
    'invoice_id':            fields.Integer(description='Invoice ID'),
    'invoice_number':        fields.String(description='Invoice number'),
    'original_order_id':     fields.String(description='Original order ID'),
    'customer_name':         fields.String(description='Customer name'),
    'invoice_date':          fields.String(description='Invoice date'),
    'total_invoice_amount':  fields.String(description='Total amount'),
    'invoice_status':        fields.String(description='Invoice status'),
    'part_no':               fields.String(description='Part number'),
    'part_name':             fields.String(description='Part name'),
    'quantity':              fields.Integer(description='Quantity'),
    'unit_price':            fields.String(description='Unit price'),
})

invoice_list_response = rest_api.model('InvoiceListResponse', {
    'success':     fields.Boolean(description='Success status'),
    'invoices':    fields.List(fields.Nested(invoice_model), description='List of invoices'),
    'total_count': fields.Integer(description='Total number of invoices'),
    'page':        fields.Integer(description='Current page'),
    'per_page':    fields.Integer(description='Items per page'),
})

invoice_error_response = rest_api.model('InvoiceErrorResponse', {
    'success': fields.Boolean(description='Success status'),
    'msg':     fields.String(description='Error message'),
})

# ── Endpoints ────────────────────────────────────────────────────────────────

@rest_api.route('/api/invoices/upload')
class InvoiceUpload(Resource):
    """MySQL handles the upload of Excel/CSV files containing invoice data."""

    @rest_api.expect(invoice_upload_parser)
    @rest_api.response(200, 'Success', invoice_upload_response)
    @rest_api.response(400, 'Bad Request', invoice_upload_response)
    @token_required
    @active_required
    @upload_permission_required('invoices')
    def post(self, current_user):
        try:
            args = invoice_upload_parser.parse_args()
            uploaded_file = args['file']
            warehouse_id  = args['warehouse_id']
            company_id    = args['company_id']

            from ..permissions import has_all_warehouse_access
            if not has_all_warehouse_access(current_user.role):
                from ..models import UserWarehouseCompany
                if not UserWarehouseCompany.user_can_access(current_user.id, warehouse_id, company_id):
                    return {'success': False, 'msg': 'You do not have access to this warehouse/company combination.',
                            'processed_count': 0, 'error_count': 0}, 403

            warehouse = Warehouse.get_by_id(warehouse_id)
            if not warehouse:
                return {'success': False, 'msg': f'Warehouse with ID {warehouse_id} not found',
                        'processed_count': 0, 'error_count': 0}, 400

            company = Company.get_by_id(company_id)
            if not company:
                return {'success': False, 'msg': f'Company with ID {company_id} not found',
                        'processed_count': 0, 'error_count': 0}, 400

            return invoice_service.process_invoice_upload(uploaded_file, warehouse_id, company_id, current_user.id)

        except Exception as e:
            return {'success': False, 'msg': f'Error processing upload: {str(e)}',
                    'processed_count': 0, 'error_count': 0}, 400


@rest_api.route('/api/invoices/download-errors')
class InvoiceErrorDownload(Resource):
    """Download error CSV from invoice upload."""

    @token_required
    @active_required
    def post(self, _current_user):
        try:
            error_csv_content = request.json.get('error_csv_content', '')

            if not error_csv_content:
                return {'success': False, 'msg': 'No error data available'}, 400

            response = make_response(error_csv_content)
            response.headers['Content-Type'] = 'text/csv'
            response.headers['Content-Disposition'] = 'attachment; filename=invoice_upload_errors.csv'
            return response

        except Exception as e:
            return {'success': False, 'msg': f'Error creating error file: {str(e)}'}, 400


@rest_api.route('/api/invoices/statistics')
class InvoiceStatistics(Resource):
    """Get invoice upload statistics."""

    @rest_api.response(200, 'Success', invoice_statistics_response)
    @rest_api.response(400, 'Error', invoice_error_response)
    @token_required
    @active_required
    def get(self, _current_user):
        """Get invoice statistics."""
        try:
            warehouse_id = request.args.get('warehouse_id', type=int)
            company_id   = request.args.get('company_id',   type=int)
            batch_id     = request.args.get('batch_id')

            stats = invoice_service.get_invoice_statistics(
                warehouse_id=warehouse_id,
                company_id=company_id,
                batch_id=batch_id
            )

            return {'success': True, **stats}, 200

        except Exception as e:
            return {'success': False, 'msg': f'Error retrieving statistics: {str(e)}'}, 400


@rest_api.route('/api/invoices')
class InvoiceList(Resource):
    """Get list of invoices with pagination and filtering."""

    @rest_api.response(200, 'Success', invoice_list_response)
    @rest_api.response(400, 'Error', invoice_error_response)
    @token_required
    @active_required
    def get(self, _current_user):
        """Get list of invoices."""
        try:
            warehouse_id = request.args.get('warehouse_id', type=int)
            company_id   = request.args.get('company_id',   type=int)
            batch_id     = request.args.get('batch_id')
            page         = request.args.get('page',     1,  type=int)
            per_page     = request.args.get('per_page', 50, type=int)

            per_page = min(per_page, 100)

            pf_sql, pf_params = partition_filter('invoice')
            base_query  = f"SELECT * FROM invoice WHERE {pf_sql}"
            count_query = f"SELECT COUNT(*) as count FROM invoice WHERE {pf_sql}"
            params = list(pf_params)

            if warehouse_id:
                base_query  += " AND warehouse_id = %s"
                count_query += " AND warehouse_id = %s"
                params.append(warehouse_id)
            if company_id:
                base_query  += " AND company_id = %s"
                count_query += " AND company_id = %s"
                params.append(company_id)
            if batch_id:
                base_query  += " AND upload_batch_id = %s"
                count_query += " AND upload_batch_id = %s"
                params.append(batch_id)

            total_result = mysql_manager.execute_query(count_query, params)
            total_count  = total_result[0]['count'] if total_result else 0

            base_query += " ORDER BY created_at DESC LIMIT %s OFFSET %s"
            params.extend([per_page, (page - 1) * per_page])

            invoices_data = mysql_manager.execute_query(base_query, params)

            invoice_list = []
            for invoice_data in invoices_data:
                invoice = Invoice(**invoice_data)
                invoice_dict = {
                    'invoice_id':           invoice.invoice_id,
                    'invoice_number':       invoice.invoice_number,
                    'original_order_id':    invoice.original_order_id,
                    'customer_name':        invoice.customer_name,
                    'invoice_date':         invoice.invoice_date.isoformat() if invoice.invoice_date else None,
                    'total_invoice_amount': str(invoice.total_invoice_amount) if invoice.total_invoice_amount else None,
                    'invoice_status':       invoice.invoice_status,
                    'part_no':              invoice.part_no,
                    'part_name':            invoice.part_name,
                    'quantity':             invoice.quantity,
                    'unit_price':           str(invoice.unit_price) if invoice.unit_price else None,
                }
                invoice_list.append(invoice_dict)

            return {
                'success':     True,
                'invoices':    invoice_list,
                'total_count': total_count,
                'page':        page,
                'per_page':    per_page,
            }, 200

        except Exception as e:
            return {'success': False, 'msg': f'Error retrieving invoices: {str(e)}'}, 400


@rest_api.route('/api/invoices/<int:invoice_id>')
class InvoiceDetail(Resource):
    """Get detailed information about a specific invoice."""

    @rest_api.response(200, 'Success')
    @rest_api.response(404, 'Invoice not found', invoice_error_response)
    def get(self, invoice_id):
        """Get invoice details."""
        try:
            invoice = Invoice.get_by_id(invoice_id)

            if not invoice:
                return {'success': False, 'msg': 'Invoice not found'}, 404

            order_info = None
            if invoice.potential_order_id:
                order = PotentialOrder.get_by_id(invoice.potential_order_id)
                if order:
                    order_info = {
                        'order_request_id': f"PO{order.potential_order_id}",
                        'status':           order.status,
                        'order_date':       order.order_date.isoformat() if order.order_date else None,
                    }

            invoice_dict = {
                'invoice_id':           invoice.invoice_id,
                'invoice_number':       invoice.invoice_number,
                'original_order_id':    invoice.original_order_id,
                'customer_name':        invoice.customer_name,
                'invoice_date':         invoice.invoice_date.isoformat() if invoice.invoice_date else None,
                'total_invoice_amount': str(invoice.total_invoice_amount) if invoice.total_invoice_amount else None,
                'invoice_status':       invoice.invoice_status,
                'part_no':              invoice.part_no,
                'part_name':            invoice.part_name,
                'quantity':             invoice.quantity,
                'unit_price':           str(invoice.unit_price) if invoice.unit_price else None,
                'order_info':           order_info,
            }

            return {'success': True, 'invoice': invoice_dict}, 200

        except Exception as e:
            return {'success': False, 'msg': f'Error retrieving invoice details: {str(e)}'}, 400


@rest_api.route('/api/invoices/supply-sheet/download')
class SupplySheetDownload(Resource):
    """Download Supply Sheet as Excel or PNG matching the exact format from the image."""

    @rest_api.doc(params={
        'format':       'Download format (excel or png)',
        'warehouse_id': 'Warehouse ID',
        'company_id':   'Company ID',
        'start_date':   'Start date (YYYY-MM-DD)',
        'end_date':     'End date (YYYY-MM-DD)',
        'batch_id':     'Upload batch ID (optional)',
    })
    def get(self):
        """Download supply sheet in specified format."""
        try:
            format_type  = request.args.get('format', 'excel').lower()
            warehouse_id = request.args.get('warehouse_id', type=int)
            company_id   = request.args.get('company_id',   type=int)
            start_date   = request.args.get('start_date')
            end_date     = request.args.get('end_date')
            batch_id     = request.args.get('batch_id')

            pf_sql, pf_params = partition_filter('invoice')
            query = f"""
                SELECT
                    invoice_number,
                    original_order_id,
                    customer_name,
                    total_invoice_amount,
                    quantity,
                    invoice_date,
                    part_no,
                    part_name
                FROM invoice
                WHERE {pf_sql}
            """
            params = list(pf_params)

            if warehouse_id:
                query += " AND warehouse_id = %s"
                params.append(warehouse_id)
            if company_id:
                query += " AND company_id = %s"
                params.append(company_id)
            if start_date:
                query += " AND DATE(invoice_date) >= %s"
                params.append(start_date)
            if end_date:
                query += " AND DATE(invoice_date) <= %s"
                params.append(end_date)
            if batch_id:
                query += " AND upload_batch_id = %s"
                params.append(batch_id)

            query += " ORDER BY invoice_date DESC, invoice_number"
            invoices_data = mysql_manager.execute_query(query, params)

            if format_type == 'excel':
                return self._generate_excel(invoices_data)
            elif format_type == 'png':
                return self._generate_png(invoices_data)
            else:
                return {'success': False, 'msg': 'Invalid format. Use "excel" or "png"'}, 400

        except Exception as e:
            return {'success': False, 'msg': f'Error generating supply sheet: {str(e)}'}, 400

    def _generate_excel(self, invoices_data):
        """Generate Excel file matching the image format exactly."""
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Supply Sheet"

        ws.merge_cells('A1:H1')
        ws['A1'] = 'On Marketing - Supply Sheet'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:H2')
        current_date = datetime.now().strftime('%d/%m/%Y')
        ws['A2'] = f'Date: {current_date}'
        ws['A2'].alignment = Alignment(horizontal='center')

        headers = ['Invoice Number', 'Order No.', 'Agencies Name', 'Invoice Value',
                   'Cases', 'Part No', 'Type', 'CODE']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'),  bottom=Side(style='thin')
            )
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        row_num     = 5
        total_value = 0
        total_cases = 0

        for invoice in invoices_data:
            invoice_number = invoice.get('invoice_number') or ''
            order_number   = invoice.get('original_order_id') or ''
            agency_name    = invoice.get('customer_name') or ''
            invoice_value  = float(invoice.get('total_invoice_amount') or 0)
            cases          = int(invoice.get('quantity') or 0)
            part_no        = invoice.get('part_no') or ''
            type_value     = 'STD' if part_no else 'GEN'
            code           = '3' if invoice_value > 1000 else '1'

            data_row = [invoice_number, order_number, agency_name, invoice_value,
                        cases, part_no, type_value, code]
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if col == 4:
                    cell.number_format = '#,##0.00'

            total_value += invoice_value
            total_cases += cases
            row_num += 1

        total_row = row_num
        ws.cell(row=total_row, column=3, value='TOTAL').font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=total_value).font = Font(bold=True)
        ws.cell(row=total_row, column=4).number_format = '#,##0.00'
        ws.cell(row=total_row, column=5, value=total_cases).font = Font(bold=True)

        for col in range(1, 9):
            cell = ws.cell(row=total_row, column=col)
            cell.border = Border(
                left=Side(style='thin'),  right=Side(style='thin'),
                top=Side(style='thick'),  bottom=Side(style='thick')
            )

        column_widths = [15, 12, 25, 12, 8, 12, 8, 8]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = (
            f'attachment; filename=supply_sheet_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        return response

    def _generate_png(self, invoices_data):
        """Generate PNG image matching the image format exactly."""
        try:
            from io import BytesIO
            from PIL import Image, ImageDraw, ImageFont

            img_width    = 1200
            header_height = 100
            row_height   = 25
            total_rows   = len(invoices_data) + 1
            img_height   = header_height + (total_rows * row_height) + 50

            img  = Image.new('RGB', (img_width, img_height), color='white')
            draw = ImageDraw.Draw(img)

            try:
                title_font  = ImageFont.truetype("arial.ttf", 16)
                header_font = ImageFont.truetype("arial.ttf", 12)
                data_font   = ImageFont.truetype("arial.ttf", 10)
            except Exception:
                title_font  = ImageFont.load_default()
                header_font = ImageFont.load_default()
                data_font   = ImageFont.load_default()

            title = "On Marketing - Supply Sheet"
            title_bbox  = draw.textbbox((0, 0), title, font=title_font)
            title_width = title_bbox[2] - title_bbox[0]
            draw.text(((img_width - title_width) // 2, 20), title, fill='black', font=title_font)

            current_date = datetime.now().strftime('%d/%m/%Y')
            date_text  = f"Date: {current_date}"
            date_bbox  = draw.textbbox((0, 0), date_text, font=header_font)
            date_width = date_bbox[2] - date_bbox[0]
            draw.text(((img_width - date_width) // 2, 50), date_text, fill='black', font=header_font)

            columns = [
                {'name': 'Invoice Number', 'width': 150, 'x': 50},
                {'name': 'Order No.',      'width': 120, 'x': 200},
                {'name': 'Agencies Name',  'width': 250, 'x': 320},
                {'name': 'Invoice Value',  'width': 120, 'x': 570},
                {'name': 'Cases',          'width': 80,  'x': 690},
                {'name': 'Part No',        'width': 100, 'x': 770},
                {'name': 'Type',           'width': 60,  'x': 870},
                {'name': 'CODE',           'width': 60,  'x': 930},
            ]

            y_start = header_height
            draw.rectangle([40, y_start, img_width - 40, y_start + row_height],
                           fill='lightgray', outline='black')
            for col in columns:
                draw.text((col['x'], y_start + 5), col['name'], fill='black', font=header_font)

            y_current   = y_start + row_height
            total_value = 0
            total_cases = 0

            for invoice in invoices_data:
                if (y_current - y_start) // row_height % 2 == 0:
                    draw.rectangle([40, y_current, img_width - 40, y_current + row_height],
                                   fill='#f8f8f8', outline='black', width=1)
                else:
                    draw.rectangle([40, y_current, img_width - 40, y_current + row_height],
                                   fill='white', outline='black', width=1)

                invoice_number = str(invoice.get('invoice_number') or '')
                order_number   = str(invoice.get('original_order_id') or '')
                agency_name    = str(invoice.get('customer_name') or '')[:25]
                invoice_value  = float(invoice.get('total_invoice_amount') or 0)
                cases          = int(invoice.get('quantity') or 0)
                part_no        = str(invoice.get('part_no') or '')
                type_value     = 'STD' if part_no else 'GEN'
                code           = '3' if invoice_value > 1000 else '1'

                data_values = [invoice_number, order_number, agency_name,
                               f"{invoice_value:,.2f}", str(cases), part_no, type_value, code]
                for col, value in zip(columns, data_values):
                    draw.text((col['x'], y_current + 5), value, fill='black', font=data_font)

                total_value += invoice_value
                total_cases += cases
                y_current   += row_height

            draw.rectangle([40, y_current, img_width - 40, y_current + row_height],
                           fill='lightblue', outline='black', width=2)
            draw.text((columns[2]['x'], y_current + 5), "TOTAL", fill='black', font=header_font)
            draw.text((columns[3]['x'], y_current + 5), f"{total_value:,.2f}", fill='black', font=header_font)
            draw.text((columns[4]['x'], y_current + 5), str(total_cases), fill='black', font=header_font)

            for col in columns[1:]:
                draw.line([col['x'] - 10, y_start, col['x'] - 10, y_current + row_height],
                          fill='black', width=1)

            output = BytesIO()
            img.save(output, format='PNG')
            output.seek(0)

            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'image/png'
            response.headers['Content-Disposition'] = (
                f'attachment; filename=supply_sheet_{datetime.now().strftime("%Y%m%d_%H%M%S")}.png'
            )
            return response

        except Exception as e:
            return {'success': False, 'msg': f'Error generating PNG: {str(e)}'}, 400
