# -*- encoding: utf-8 -*-

from flask import request, send_file
from flask_restx import Resource, fields
from datetime import datetime, timedelta
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from .models import (
    Users, JWTTokenBlocklist, Warehouse, Company, PotentialOrder,
    PotentialOrderProduct, OrderStateHistory, OrderState, Product,
    Dealer, Box, Order, OrderProduct, OrderBox, BoxProduct, Invoice,
    UserWarehouseCompany, mysql_manager
)
from .routes import rest_api, token_required, active_required
from .permissions import get_permissions, has_all_warehouse_access

# Response models
warehouse_model = rest_api.model('Warehouse', {
    'id': fields.Integer(description='Warehouse ID'),
    'name': fields.String(description='Warehouse name'),
    'location': fields.String(description='Warehouse location')
})

warehouses_response = rest_api.model('WarehousesResponse', {
    'success': fields.Boolean(description='Success status'),
    'warehouses': fields.List(fields.Nested(warehouse_model), description='List of warehouses')
})

company_model = rest_api.model('Company', {
    'id': fields.Integer(description='Company ID'),
    'name': fields.String(description='Company name')
})

companies_response = rest_api.model('CompaniesResponse', {
    'success': fields.Boolean(description='Success status'),
    'companies': fields.List(fields.Nested(company_model), description='List of companies')
})

status_count_model = rest_api.model('StatusCount', {
    'count': fields.Integer(description='Number of orders with this status'),
    'label': fields.String(description='Label for the status')
})

status_counts_model = rest_api.model('StatusCounts', {
    'open': fields.Nested(status_count_model),
    'picking': fields.Nested(status_count_model),
    'packed': fields.Nested(status_count_model),
    'invoiced': fields.Nested(status_count_model),
    'dispatch-ready': fields.Nested(status_count_model),
    'completed': fields.Nested(status_count_model),
    'partially-completed': fields.Nested(status_count_model)
})

status_response = rest_api.model('StatusResponse', {
    'success': fields.Boolean(description='Success status'),
    'status_counts': fields.Nested(status_counts_model, description='Order status counts')
})

state_history_model = rest_api.model('StateHistory', {
    'state_name': fields.String(description='State name'),
    'timestamp': fields.String(description='Timestamp of state change'),
    'user': fields.String(description='User who changed the state')
})

order_model = rest_api.model('Order', {
    'order_request_id': fields.String(description='Order request ID'),
    'original_order_id': fields.String(description='Original order ID'),
    'dealer_name': fields.String(description='Dealer name'),
    'order_date': fields.String(description='Order date'),
    'status': fields.String(description='Current status'),
    'current_state_time': fields.String(description='Time of current state'),
    'assigned_to': fields.String(description='User assigned to this order'),
    'products': fields.Integer(description='Number of products in this order'),
    'state_history': fields.List(fields.Nested(state_history_model))
})

orders_response = rest_api.model('OrdersResponse', {
    'success': fields.Boolean(description='Success status'),
    'orders': fields.List(fields.Nested(order_model), description='List of orders')
})

order_detail_response = rest_api.model('OrderDetailResponse', {
    'success': fields.Boolean(description='Success status'),
    'order': fields.Nested(order_model, description='Order details')
})

recent_order_model = rest_api.model('RecentOrder', {
    'order_request_id': fields.String(description='Order request ID'),
    'dealer_name': fields.String(description='Dealer name'),
    'status': fields.String(description='Current status'),
    'order_date': fields.String(description='Order date'),
    'current_state_time': fields.String(description='Time of current state'),
    'assigned_to': fields.String(description='User assigned to this order')
})

recent_orders_response = rest_api.model('RecentOrdersResponse', {
    'success': fields.Boolean(description='Success status'),
    'recent_orders': fields.List(fields.Nested(recent_order_model), description='List of recent orders')
})

error_response = rest_api.model('ErrorResponse', {
    'success': fields.Boolean(description='Success status'),
    'msg': fields.String(description='Error message')
})


@rest_api.route('/api/warehouses')
class WarehouseList(Resource):
    @rest_api.marshal_with(warehouses_response)
    @rest_api.response(400, 'Error', error_response)
    @token_required
    @active_required
    def get(self, current_user):
        try:
            warehouses = Warehouse.get_all()
            warehouse_list = [{'id': w.warehouse_id, 'name': w.name, 'location': w.location} for w in warehouses]
            return {'success': True, 'warehouses': warehouse_list}, 200
        except Exception as e:
            return {'success': False, 'msg': f'Error retrieving warehouses: {str(e)}'}, 400


@rest_api.route('/api/companies')
class CompanyList(Resource):
    @rest_api.marshal_with(companies_response)
    @rest_api.response(400, 'Error', error_response)
    @token_required
    @active_required
    def get(self, current_user):
        try:
            warehouse_id = request.args.get('warehouse_id', type=int)

            companies = Company.get_all()
            company_list = [{'id': c.company_id, 'name': c.name} for c in companies]
            return {'success': True, 'companies': company_list}, 200
        except Exception as e:
            return {'success': False, 'msg': f'Error retrieving companies: {str(e)}'}, 400


@rest_api.route('/api/orders/status')
class OrderStatusCount(Resource):
    """
    MySQL endpoint for retrieving order status counts
    """

    @rest_api.doc(params={'warehouse_id': 'Warehouse ID', 'company_id': 'Company ID'})
    @rest_api.marshal_with(status_response)
    @rest_api.response(400, 'Error', error_response)
    @token_required
    @active_required
    def get(self, current_user):
        try:
            warehouse_id = request.args.get('warehouse_id', type=int)
            company_id = request.args.get('company_id', type=int)

            all_statuses = [
                ('open',                'Open',               'Open Orders'),
                ('picking',             'Picking',            'Picking'),
                ('packed',              'Packed',             'Packed'),
                ('invoiced',            'Invoiced',           'Invoiced'),
                ('dispatch-ready',      'Dispatch Ready',     'Dispatch Ready'),
                ('completed',           'Completed',          'Completed'),
                ('partially-completed', 'Partially Completed','Partially Completed'),
            ]

            response_data = {}
            for key, db_status, label in all_statuses:
                count = PotentialOrder.count_by_status(db_status, warehouse_id, company_id)
                response_data[key] = {'count': count, 'label': label}

            return {'success': True, 'status_counts': response_data}, 200

        except Exception as e:
            return {'success': False, 'msg': f'Error retrieving order status counts: {str(e)}'}, 400

@rest_api.route('/api/orders')
class OrdersList(Resource):
    """
    MySQL endpoint for retrieving orders filtered by status
    """

    @rest_api.doc(params={
        'status': 'Order status (open, picking, packed, dispatch-ready, completed, partially-completed)',
        'warehouse_id': 'Warehouse ID',
        'company_id': 'Company ID',
        'limit': 'Limit number of results (default 100)'
    })
    @rest_api.marshal_with(orders_response)
    @rest_api.response(400, 'Error', error_response)
    @token_required
    @active_required
    def get(self, current_user):
        try:
            status = request.args.get('status', '')
            warehouse_id = request.args.get('warehouse_id', type=int)
            company_id = request.args.get('company_id', type=int)
            limit = request.args.get('limit', 1000, type=int)

            # Map frontend status names to database status names
            status_map = {
                'open': 'Open',
                'picking': 'Picking',
                'packed': 'Packed',
                'invoiced': 'Invoiced',
                'dispatch-ready': 'Dispatch Ready',
                'completed': 'Completed',
                'partially-completed': 'Partially Completed'
            }

            db_status = status_map.get(status.lower(), '') if status else ''

            potential_orders = PotentialOrder.find_by_filters(
                status=db_status,
                warehouse_id=warehouse_id,
                company_id=company_id,
                limit=limit
            )

            frontend_status_map = {
                'Open': 'open',
                'Picking': 'picking',
                'Packed': 'packed',
                'Invoiced': 'invoiced',
                'Dispatch Ready': 'dispatch-ready',
                'Completed': 'completed',
                'Partially Completed': 'partially-completed'
            }

            orders = []
            for order_data in potential_orders:
                try:
                    dealer_name = order_data.get('dealer_name') or 'Unknown Dealer'
                    product_count = PotentialOrderProduct.count_by_order(order_data['potential_order_id'])
                    state_history_data = OrderStateHistory.get_history_for_order(order_data['potential_order_id'])

                    formatted_history = []
                    for history in state_history_data:
                        formatted_history.append({
                            'state_name': history['state_name'],
                            'timestamp': history['changed_at'].isoformat() if history['changed_at'] else None,
                            'user': f"User {history['changed_by']}"
                        })

                    # Null-safe current state time
                    current_state_time = (
                        state_history_data[-1]['changed_at'] if state_history_data
                        else order_data.get('updated_at') or order_data.get('created_at')
                    )
                    current_state_time_str = current_state_time.isoformat() if current_state_time else None

                    # Null-safe order date
                    order_date = order_data.get('order_date')
                    order_date_str = order_date.isoformat() if order_date else None

                    db_status = order_data.get('status', '')
                    frontend_status = frontend_status_map.get(db_status, db_status.lower().replace(' ', '-'))

                    orders.append({
                        'order_request_id': f"PO{order_data['potential_order_id']}",
                        'original_order_id': order_data['original_order_id'],
                        'dealer_name': dealer_name,
                        'order_date': order_date_str,
                        'status': frontend_status,
                        'current_state_time': current_state_time_str,
                        'assigned_to': f"User {order_data['requested_by']}",
                        'products': product_count,
                        'state_history': formatted_history,
                        'invoice_submitted': bool(order_data.get('invoice_submitted', False)),
                    })
                except Exception as row_err:
                    print(f"Warning: skipping order {order_data.get('potential_order_id')} due to error: {row_err}")

            return {
                'success': True,
                'orders': orders
            }, 200

        except Exception as e:
            print(f"Error in /api/orders: {str(e)}")
            import traceback
            traceback.print_exc()
            return {
                'success': False,
                'msg': f'Error retrieving orders: {str(e)}'
            }, 400

FRONTEND_TO_DB_STATUS = {
    'open': 'Open',
    'picking': 'Picking',
    'packed': 'Packed',
    'invoiced': 'Invoiced',
    'dispatch-ready': 'Dispatch Ready',
    'completed': 'Completed',
}

DB_TO_FRONTEND_STATUS = {v: k for k, v in FRONTEND_TO_DB_STATUS.items()}

# open→picking, picking→packed, packed→invoiced, dispatch-ready→completed
VALID_BULK_TRANSITIONS = {
    'open': 'picking',
    'picking': 'packed',
    'packed': 'invoiced',
    'dispatch-ready': 'completed',
}


@rest_api.route('/api/orders/bulk-export')
class BulkOrderExport(Resource):
    """Download orders as Excel template for bulk status update"""

    @token_required
    @active_required
    def get(self, current_user):
        """Generate and return Excel file for bulk order status update"""
        try:
            status = request.args.get('status', '')
            warehouse_id = request.args.get('warehouse_id', type=int)
            company_id = request.args.get('company_id', type=int)

            db_status = FRONTEND_TO_DB_STATUS.get(status.lower(), '') if status else ''

            potential_orders = PotentialOrder.find_by_filters(
                status=db_status,
                warehouse_id=warehouse_id,
                company_id=company_id,
                limit=1000
            )

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Orders'

            # --- Header row ---
            header_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            headers = ['Order ID', 'Customer Name', 'Current Status', 'Expected Status', 'Number of Boxes']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # --- Instructions row ---
            note_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
            note_font = Font(italic=True, color='5D4037')
            notes = [
                '(do not edit)',
                '(do not edit)',
                '(do not edit)',
                'Fill: picking / packed / invoiced / completed',
                'Fill only when moving packed → invoiced',
            ]
            for col, note in enumerate(notes, 1):
                cell = ws.cell(row=2, column=col, value=note)
                cell.fill = note_fill
                cell.font = note_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # --- Blocked note row ---
            blocked_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
            ws.cell(row=3, column=1, value='NOTE').fill = blocked_fill
            note_cell = ws.cell(
                row=3, column=2,
                value='invoiced → dispatch-ready is NOT allowed here. Use the Invoice Upload tab.'
            )
            note_cell.fill = blocked_fill
            note_cell.font = Font(bold=True, color='B71C1C')
            ws.merge_cells('B3:E3')

            # --- Data rows (starting row 4) ---
            readonly_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
            for row_idx, order_data in enumerate(potential_orders, 4):
                current_fe_status = DB_TO_FRONTEND_STATUS.get(order_data['status'],
                                                               order_data['status'].lower().replace(' ', '-'))
                dealer_name = order_data.get('dealer_name', 'Unknown')

                for col, val in enumerate([
                    f"PO{order_data['potential_order_id']}",
                    dealer_name,
                    current_fe_status,
                    '',   # Expected Status — user fills
                    '',   # Number of Boxes — user fills if needed
                ], 1):
                    cell = ws.cell(row=row_idx, column=col, value=val)
                    if col <= 3:
                        cell.fill = readonly_fill

            # --- Column widths ---
            ws.column_dimensions['A'].width = 14
            ws.column_dimensions['B'].width = 32
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 22
            ws.row_dimensions[1].height = 20
            ws.row_dimensions[2].height = 18

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            filename = f'orders_bulk_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )

        except Exception as e:
            import traceback
            traceback.print_exc()
            return {'success': False, 'msg': f'Error generating export: {str(e)}'}, 400


@rest_api.route('/api/orders/bulk-import')
class BulkOrderImport(Resource):
    """Upload filled Excel template to perform bulk order status transitions"""

    @token_required
    @active_required
    def post(self, current_user):
        """Process bulk order status updates from uploaded Excel"""
        try:
            if 'file' not in request.files:
                return {'success': False, 'msg': 'No file uploaded'}, 400

            file = request.files['file']
            try:
                wb = openpyxl.load_workbook(io.BytesIO(file.read()))
                ws = wb.active
            except Exception as e:
                return {'success': False, 'msg': f'Invalid Excel file: {str(e)}'}, 400

            moved, skipped, errors = [], [], []

            # Data starts at row 4 (rows 1-3 are header / notes)
            for row in ws.iter_rows(min_row=4, values_only=True):
                order_id = str(row[0] or '').strip() if row[0] else ''
                current_status = str(row[2] or '').strip().lower()
                expected_status = str(row[3] or '').strip().lower()
                raw_boxes = row[4]

                if not order_id:
                    continue

                if not expected_status:
                    skipped.append({'order_id': order_id, 'reason': 'No expected status provided'})
                    continue

                if current_status == expected_status:
                    skipped.append({'order_id': order_id, 'reason': 'Status unchanged'})
                    continue

                # Block invoiced → dispatch-ready
                if current_status == 'invoiced' and expected_status == 'dispatch-ready':
                    errors.append({
                        'order_id': order_id,
                        'reason': 'invoiced → dispatch-ready is not allowed here. Use the Invoice Upload tab.'
                    })
                    continue

                # Validate transition
                if VALID_BULK_TRANSITIONS.get(current_status) != expected_status:
                    valid_next = VALID_BULK_TRANSITIONS.get(current_status, 'none')
                    errors.append({
                        'order_id': order_id,
                        'reason': f'Invalid transition: {current_status} → {expected_status}. Valid next: {valid_next}'
                    })
                    continue

                # Parse order numeric ID
                try:
                    numeric_id = int(order_id.replace('PO', ''))
                except ValueError:
                    errors.append({'order_id': order_id, 'reason': 'Invalid order ID format'})
                    continue

                potential_order = PotentialOrder.get_by_id(numeric_id)
                if not potential_order:
                    errors.append({'order_id': order_id, 'reason': 'Order not found'})
                    continue

                # Verify current status matches DB
                db_current = DB_TO_FRONTEND_STATUS.get(potential_order.status,
                                                        potential_order.status.lower().replace(' ', '-'))
                if db_current != current_status:
                    errors.append({
                        'order_id': order_id,
                        'reason': f'Status mismatch: DB has "{db_current}", Excel shows "{current_status}"'
                    })
                    continue

                try:
                    current_time = datetime.utcnow()

                    if expected_status == 'invoiced':
                        # Packed → Invoiced
                        try:
                            number_of_boxes = max(1, int(raw_boxes or 1))
                        except (ValueError, TypeError):
                            number_of_boxes = 1

                        final_order = Order(
                            potential_order_id=numeric_id,
                            order_number=f"ORD-{numeric_id}-{current_time.strftime('%Y%m%d%H%M')}",
                            status='Dispatch Ready',
                            created_at=current_time,
                            updated_at=current_time
                        )
                        final_order.save()

                        for i in range(number_of_boxes):
                            OrderBox(
                                order_id=final_order.order_id,
                                name=f'Box-{i + 1}',
                                created_at=current_time,
                                updated_at=current_time
                            ).save()

                        potential_order.status = 'Invoiced'
                        potential_order.updated_at = current_time
                        potential_order.save()

                        state = OrderState.find_by_name('Invoiced') or _ensure_state('Invoiced', 'Order invoiced and ready for dispatch')
                        OrderStateHistory(
                            potential_order_id=numeric_id,
                            state_id=state.state_id,
                            changed_by=current_user.id,
                            changed_at=current_time
                        ).save()

                        moved.append({
                            'order_id': order_id,
                            'from': current_status,
                            'to': expected_status,
                            'boxes': number_of_boxes
                        })

                    elif expected_status == 'completed':
                        # Dispatch Ready → Completed
                        final_order = Order.find_by_potential_order_id(numeric_id)
                        if not final_order:
                            errors.append({'order_id': order_id, 'reason': 'No final order found for dispatch-ready order'})
                            continue

                        final_order.status = 'Completed'
                        final_order.dispatched_date = current_time
                        final_order.updated_at = current_time
                        final_order.save()

                        potential_order.status = 'Completed'
                        potential_order.updated_at = current_time
                        potential_order.save()

                        state = OrderState.find_by_name('Completed') or _ensure_state('Completed', 'Order completed and dispatched')
                        OrderStateHistory(
                            potential_order_id=numeric_id,
                            state_id=state.state_id,
                            changed_by=current_user.id,
                            changed_at=current_time
                        ).save()

                        moved.append({'order_id': order_id, 'from': current_status, 'to': expected_status})

                    else:
                        # Simple transitions: open→picking, picking→packed
                        new_db_status = FRONTEND_TO_DB_STATUS[expected_status]
                        potential_order.status = new_db_status
                        potential_order.updated_at = current_time
                        potential_order.save()

                        state = OrderState.find_by_name(new_db_status) or _ensure_state(new_db_status, f'Order moved to {new_db_status}')
                        OrderStateHistory(
                            potential_order_id=numeric_id,
                            state_id=state.state_id,
                            changed_by=current_user.id,
                            changed_at=current_time
                        ).save()

                        moved.append({'order_id': order_id, 'from': current_status, 'to': expected_status})

                except Exception as e:
                    errors.append({'order_id': order_id, 'reason': f'Error processing: {str(e)}'})

            return {
                'success': True,
                'summary': {
                    'moved': len(moved),
                    'skipped': len(skipped),
                    'errors': len(errors)
                },
                'details': {'moved': moved, 'skipped': skipped, 'errors': errors}
            }, 200

        except Exception as e:
            import traceback
            traceback.print_exc()
            return {'success': False, 'msg': f'Error processing import: {str(e)}'}, 400


def _ensure_state(name, description):
    """Get or create an OrderState by name."""
    state = OrderState.find_by_name(name)
    if not state:
        state = OrderState(state_name=name, description=description)
        state.save()
    return state


@rest_api.route('/api/orders/<string:order_id>')
class OrderDetail(Resource):
    """
    MySQL endpoint for retrieving details of a specific order
    """

    @rest_api.marshal_with(order_detail_response)
    @rest_api.response(400, 'Error', error_response)
    @rest_api.response(404, 'Order not found', error_response)
    def get(self, order_id):
        """Get order details - MySQL implementation"""
        try:
            # Extract the numeric part from the order_id
            numeric_id = int(order_id.replace('PO', '')) if order_id.startswith('PO') else int(order_id)

            # Get the order using MySQL
            potential_order = PotentialOrder.get_by_id(numeric_id)
            if not potential_order:
                return {
                    'success': False,
                    'msg': 'Order not found'
                }, 404

            # Get dealer information
            dealer = None
            if potential_order.dealer_id:
                dealer = Dealer.get_by_id(potential_order.dealer_id)
            dealer_name = dealer.name if dealer else 'Unknown Dealer'

            # Get product count
            product_count = PotentialOrderProduct.count_by_order(numeric_id)

            # Get state history
            state_history_data = OrderStateHistory.get_history_for_order(numeric_id)

            # Format state history
            formatted_history = []
            for history in state_history_data:
                formatted_history.append({
                    'state_name': history['state_name'],
                    'timestamp': history['changed_at'].isoformat(),
                    'user': f"User {history['changed_by']}"
                })

            # Get the time of the most recent state change
            current_state_time = potential_order.updated_at
            if state_history_data:
                current_state_time = state_history_data[-1]['changed_at']

            # Map database status to frontend status
            status_map = {
                'Open': 'open',
                'Picking': 'picking',
                'Packed': 'packed',
                'Invoiced': 'invoiced',
                'Dispatch Ready': 'dispatch-ready',
                'Completed': 'completed',
                'Partially Completed': 'partially-completed'
            }
            status = status_map.get(potential_order.status, 'open')

            # Format the order
            order_data = {
                'order_request_id': f"PO{potential_order.potential_order_id}",
                'original_order_id': potential_order.original_order_id,
                'dealer_name': dealer_name,
                'order_date': potential_order.order_date.isoformat(),
                'status': status,
                'current_state_time': current_state_time.isoformat(),
                'assigned_to': f"User {potential_order.requested_by}",
                'products': product_count,
                'state_history': formatted_history
            }

            return {
                'success': True,
                'order': order_data
            }, 200

        except Exception as e:
            print(f"Error in /api/orders/{order_id}: {str(e)}")
            return {
                'success': False,
                'msg': f'Error retrieving order details: {str(e)}'
            }, 400

@rest_api.route('/api/orders/recent')
class RecentOrders(Resource):
    """
    MySQL endpoint for retrieving recent order activity
    """

    @rest_api.doc(params={
        'warehouse_id': 'Warehouse ID',
        'company_id': 'Company ID',
        'limit': 'Maximum number of orders to return (default 10)'
    })
    @rest_api.marshal_with(recent_orders_response)
    @rest_api.response(400, 'Error', error_response)
    @token_required
    @active_required
    def get(self, current_user):
        try:
            warehouse_id = request.args.get('warehouse_id', type=int)
            company_id = request.args.get('company_id', type=int)
            limit = request.args.get('limit', 10, type=int)

            allowed_states = get_permissions(current_user.role)['order_states']

            potential_orders = PotentialOrder.find_by_filters(
                warehouse_id=warehouse_id,
                company_id=company_id,
                limit=limit
            )
            # Only show orders in states the role can see
            potential_orders = [o for o in potential_orders if o['status'] in allowed_states]

            orders = []
            for order_data in potential_orders:
                # Get dealer name
                dealer_name = order_data.get('dealer_name', 'Unknown Dealer')

                # Get the most recent state change for each order
                state_history_data = OrderStateHistory.get_history_for_order(order_data['potential_order_id'])

                current_state_time = order_data['updated_at']
                if state_history_data:
                    current_state_time = state_history_data[-1]['changed_at']

                # Map database status to frontend status
                status_map = {
                    'Open': 'open',
                    'Picking': 'picking',
                    'Packed': 'packed',
                    'Invoiced': 'invoiced',
                    'Dispatch Ready': 'dispatch-ready',
                    'Completed': 'completed',
                    'Partially Completed': 'partially-completed'
                }
                status = status_map.get(order_data['status'], 'open')

                # Format the order
                order_result = {
                    'order_request_id': f"PO{order_data['potential_order_id']}",
                    'dealer_name': dealer_name,
                    'status': status,
                    'order_date': order_data['order_date'].isoformat(),
                    'current_state_time': current_state_time.isoformat(),
                    'assigned_to': f"User {order_data['requested_by']}",
                }
                orders.append(order_result)

            return {
                'success': True,
                'recent_orders': orders
            }, 200

        except Exception as e:
            print(f"Error in /api/orders/recent: {str(e)}")
            return {
                'success': False,
                'msg': f'Error retrieving recent orders: {str(e)}'
            }, 400